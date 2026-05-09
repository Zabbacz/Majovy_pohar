import shutil
from decimal import Decimal
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from services.VysledkyService import VysledkyService

class ExportService:

    @staticmethod
    def export_to_xlsx(source_path: str):

        data = VysledkyService.get_data()
        naradi = data["naradi"]
        kategorie_dict = data["kategorie"]

        source = Path(source_path)
        target = source.with_name(source.stem + "_vysledky.xlsx")

        shutil.copy(source, target)

        wb = load_workbook(target)

        start_col = column_index_from_string("H")
        max_col_allowed = column_index_from_string("AF")

        errors = []

        for kategorie, zavodnici in kategorie_dict.items():

            if kategorie not in wb.sheetnames:
                errors.append(f"Chybí list: {kategorie}")
                continue

            ws = wb[kategorie]

            for poradi, z in enumerate(zavodnici, start=1):

                row_found = None
                hledane_jmeno = (z["jmeno"] or "").strip()

                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=4).value
                    if cell_value and str(cell_value).strip() == hledane_jmeno:
                        row_found = row
                        break

                if not row_found:
                    errors.append(f"{kategorie} - Nenalezen: {z['jmeno']}")
                    continue

                ws.cell(row=row_found, column=1, value=poradi)

                col_offset = 0

                for nar_id, _ in naradi:

                    current_col = start_col + col_offset

                    if current_col + 3 > max_col_allowed:
                        errors.append(
                            f"{kategorie} - {z['jmeno']} přesahuje sloupec AF"
                        )
                        break

                    nar = z["naradi"].get(nar_id, {})

                    ws.cell(row=row_found, column=current_col + 0, value=nar.get("D", 0))
                    ws.cell(row=row_found, column=current_col + 1, value=nar.get("E", 0))
                    ws.cell(row=row_found, column=current_col + 2, value=nar.get("Pen", 0))
                    ws.cell(row=row_found, column=current_col + 3, value=nar.get("Vysl", 0))

                    col_offset += 4

        wb.save(target)

        return {
            "ok": len(errors) == 0,
            "file": str(target),
            "errors": errors
        }
