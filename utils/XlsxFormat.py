from decimal import Decimal
from openpyxl import load_workbook


def format_numeric_cells(file_path: str, fmt: str = '[$-409]0.000') -> None:
    """
    Přeformátuje všechny číselné buňky a buňky se vzorci v xlsx souboru
    na zadaný formát s tečkou jako desetinným oddělovačem.
    """
    wb = load_workbook(file_path)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float, Decimal)):
                    cell.value = float(cell.value)
                    cell.number_format = fmt
                elif isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.number_format = fmt

    wb.save(file_path)